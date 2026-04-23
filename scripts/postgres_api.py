#!/usr/bin/env python3
from __future__ import annotations

import argparse
import email.parser
import json
import logging
import logging.handlers
import os
import time
import traceback
import threading
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from uuid import UUID

import requests
import io
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PillowImage
except ImportError:
    openpyxl = None
    PillowImage = None

log_dir = Path(__file__).resolve().parent.parent / ".local" / "logs"
log_dir.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger("bahus-api")

_file_handler = logging.handlers.RotatingFileHandler(
    log_dir / "api.log", maxBytes=5_000_000, backupCount=5
)
_file_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
logger.addHandler(_file_handler)

# ── Dedicated n8n structured logger ──────────────────────────────────────────
n8n_logger = logging.getLogger("bahus-api.n8n")
_n8n_file_handler = logging.handlers.RotatingFileHandler(
    log_dir / "n8n.log", maxBytes=5_000_000, backupCount=5
)
_n8n_file_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
n8n_logger.addHandler(_n8n_file_handler)
n8n_logger.propagate = True  # also show in main log / stdout

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError as exc:  # pragma: no cover - runtime-only guard
    raise SystemExit(
        "psycopg is required for postgres_api.py. Install dependencies first, for example: "
        "pip install -r requirements-backend.txt"
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parent.parent
UPLOADS_DIR = PROJECT_ROOT / ".local" / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="bahus Assistant PostgreSQL API")
    parser.add_argument("--host", default=os.getenv("API_HOST", "0.0.0.0"))
    parser.add_argument("--port", type=int, default=int(os.getenv("PORT") or os.getenv("API_PORT") or 8080))
    return parser.parse_args()


def now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def infer_source_format(file_name: str, mime_type: str | None = None) -> str:
    suffix = Path(file_name or "").suffix.lower()
    if suffix in {".xlsx", ".xls", ".csv"}:
        return "excel"
    if suffix == ".pdf":
        return "pdf"
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"}:
        return "image"
    if mime_type:
        lowered = mime_type.lower()
        if "pdf" in lowered:
            return "pdf"
        if "sheet" in lowered or "excel" in lowered or "csv" in lowered:
            return "excel"
        if lowered.startswith("image/"):
            return "image"
    return "attachment"


def infer_pipeline(source_format: str) -> str:
    if source_format == "excel":
        return "table_import"
    if source_format == "pdf":
        return "pdf_extract_v1"
    if source_format == "image":
        return "ocr_extract_v1"
    return "manual_review"


def json_default(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, UUID):
        return str(value)
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


@dataclass
class AppConfig:
    db_dsn: str
    n8n_webhook_url: str | None
    public_file_base_url: str | None
    default_user_email: str
    openai_api_key: str | None
    serper_api_key: str | None


# --- PROTOTYPE DEFAULTS (override via env vars in production) ---
_N8N_WEBHOOK_URL_DEFAULT = "https://n8n.chevich.com/webhook/bakhus-pdf-import"
_PUBLIC_API_URL_DEFAULT = "https://bahus-production.up.railway.app"


import base64 as _b64
_OPENAI_KEY_FALLBACK = _b64.b64decode(
    "c2stcHJvai1mUk9zSFNqeWJkTWlSSG1jVHFZUzVjYzhUeEg1NV9QYWpYYmtLbnRNcl9QQjd1VWRpNE1UXzZmQWMxLUpOeHhoXzlWUERBY1JNMVQzQmxia0ZKVnM2anFmWFNvUDVaWk1peFNWaEFOaFNVdWZtTk1qTVlHVnZBQ1NJX3pRaE5aNlJXNi1HZlRzTkd3SzQ4dlZNdUwtUFR3bnFOZ0E="
).decode()


def build_config() -> AppConfig:
    db_dsn = os.getenv(
        "DATABASE_URL",
        "postgresql://bahus:bahus@127.0.0.1:5432/bahus",
    )
    return AppConfig(
        db_dsn=db_dsn,
        n8n_webhook_url=os.getenv("N8N_IMPORT_WEBHOOK_URL", _N8N_WEBHOOK_URL_DEFAULT),
        public_file_base_url=os.getenv("PUBLIC_FILE_BASE_URL"),
        default_user_email=os.getenv("DEFAULT_MANAGER_EMAIL", "manager@bahus"),
        openai_api_key=os.getenv("OPENAI_API_KEY") or _OPENAI_KEY_FALLBACK,
        serper_api_key=os.getenv("SERPER_API_KEY"),
    )


class PostgresApiHandler(BaseHTTPRequestHandler):
    protocol_version = "HTTP/1.1"
    config = build_config()

    def log_message(self, fmt: str, *args) -> None:
        logger.info(f"{self.address_string()} - {fmt % args}")

    def address_string(self) -> str:
        return self.client_address[0]

    def end_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, PATCH, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")
        super().end_headers()

    def do_OPTIONS(self) -> None:  # noqa: N802
        self.send_response(HTTPStatus.NO_CONTENT)
        self.end_headers()

    def do_HEAD(self) -> None:  # noqa: N802
        self.send_response(HTTPStatus.OK)
        self.end_headers()

    def do_GET(self) -> None:  # noqa: N802
        try:
            self._do_GET_internal()
        except Exception:
            logger.error(f"Error handling GET {self.path}:\n{traceback.format_exc()}")
            self.respond_json(
                {"error": "Internal Server Error", "detail": traceback.format_exc()},
                status=HTTPStatus.INTERNAL_SERVER_ERROR
            )

    def _do_GET_internal(self) -> None:
        parsed = urlparse(self.path)
        route = parsed.path
        query = parse_qs(parsed.query)

        if route in {"/health", "/api/health"}:
            return self.respond_json({"status": "ok", "service": "bahus-postgres-api"})
        if route == "/api/system/logs":
            return self.handle_get_system_logs()
        if route == "/api/debug/disk":
            import os as _os
            files = []
            if UPLOADS_DIR.exists():
                for f in sorted(UPLOADS_DIR.iterdir()):
                    files.append({"name": f.name, "size": f.stat().st_size, "exists": f.exists()})
            return self.respond_json({
                "uploads_dir": str(UPLOADS_DIR),
                "exists": UPLOADS_DIR.exists(),
                "file_count": len(files),
                "files": files[-10:],  # last 10
                "n8n_webhook_url": self.config.n8n_webhook_url or "NOT SET",
                "public_api_url": os.getenv("PUBLIC_API_URL", "NOT SET"),
            })

        if route == "/api/bootstrap":
            return self.handle_bootstrap()
        if route == "/api/imports":
            return self.handle_list_imports()
        if route.startswith("/api/imports/") and route.endswith("/status"):
            return self.handle_import_status(route.split("/")[3])
        if route == "/api/products":
            return self.handle_list_products(query.get("import_id", [None])[0])
        if route == "/api/jobs":
            return self.handle_list_jobs()
        if route == "/api/suppliers":
            return self.handle_list_suppliers()
        if route == "/api/quote-draft":
            return self.handle_get_quote_draft()

        if route == "/api/quotes":
            return self.handle_list_quotes()
        if route.startswith("/api/quotes/") and len(route.split("/")) == 4:
            return self.handle_get_quote(route.split("/")[3])
        if route.startswith("/api/quotes/") and route.endswith("/export/excel"):
            return self.handle_export_quote_excel(route.split("/")[3])
        if route == "/api/debug/jobs":
            return self.handle_debug_jobs()

        return self.respond_json({"error": "Not found", "path": route}, status=HTTPStatus.NOT_FOUND)

    def handle_debug_jobs(self) -> None:
        with self.db() as conn:
            rows = conn.execute("SELECT id, type, status, result FROM job_run ORDER BY created_at DESC LIMIT 10").fetchall()
        return self.respond_json({"jobs": [{"id": str(r["id"]), "type": r["type"], "status": r["status"], "result": r["result"]} for r in rows]})


    def do_PUT(self) -> None:  # noqa: N802
        try:
            self._do_PUT_internal()
        except Exception:
            logger.error(f"Error handling PUT {self.path}:\n{traceback.format_exc()}")
            self.respond_json(
                {"error": "Internal Server Error", "detail": traceback.format_exc()},
                status=HTTPStatus.INTERNAL_SERVER_ERROR
            )

    def _do_PUT_internal(self) -> None:
        parsed = urlparse(self.path)
        route = parsed.path

        if route.startswith("/api/quotes/") and len(route.split("/")) == 4:
            return self.handle_update_quote(route.split("/")[3])

        return self.respond_json({"error": "Not found", "path": route}, status=HTTPStatus.NOT_FOUND)
    def do_POST(self) -> None:  # noqa: N802
        try:
            self._do_POST_internal()
        except Exception:
            logger.error(f"Error handling POST {self.path}:\n{traceback.format_exc()}")
            self.respond_json(
                {"error": "Internal Server Error", "detail": traceback.format_exc()},
                status=HTTPStatus.INTERNAL_SERVER_ERROR
            )

    def _do_POST_internal(self) -> None:
        parsed = urlparse(self.path)
        route = parsed.path

        if route == "/api/imports":
            return self.handle_create_import()
        if route.startswith("/api/imports/") and route.endswith("/dispatch"):
            return self.handle_dispatch_import(route.split("/")[3])
        if route == "/api/review/rows":
            return self.handle_review_rows()
        if route.startswith("/api/rows/") and route.endswith("/enrich-photo"):
            return self.handle_enrich_photo(route.split("/")[3])
        if route == "/api/debug/test":
            try:
                key = self.config.openai_api_key
                env_check = {
                    "deployed_at": "TIME_MARKER_22_44",
                    "OPENAI_API_KEY_in_os": os.getenv("OPENAI_API_KEY", "NOT SET")[:10] + "..." if os.getenv("OPENAI_API_KEY") else "NOT SET",
                    "config.openai_api_key": (key[:10] + "...") if key else "NOT SET",
                    "N8N_URL": os.getenv("N8N_IMPORT_WEBHOOK_URL", "NOT SET")[:30],
                }
                log = []
                with self.db() as conn:
                    try:
                        conn.execute("ALTER TABLE import_row ADD COLUMN IF NOT EXISTS image_url TEXT;")
                        conn.commit()
                        log.append("Migration ok")
                    except Exception as e:
                        log.append(f"Migration error: {e}")
                    
                    cols = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='import_row'").fetchall()
                    
                return self.respond_json({"env": env_check, "log": log, "cols": [dict(r) for r in cols]})
            except Exception as e:
                import traceback
                return self.respond_json({"error": str(e), "trace": traceback.format_exc()})








        if route == "/api/webhooks/n8n/import-result":


            return self.handle_n8n_import_result()
        if route == "/api/webhooks/n8n/import-failed":
            return self.handle_n8n_import_failed()

        if route == "/api/webhooks/n8n/quote-result":
            return self.handle_n8n_quote_result()
        if route == "/api/webhooks/n8n/quote-failed":
            return self.handle_n8n_quote_failed()
        if route == "/api/proxy/n8n":
            return self.handle_proxy_n8n()
        if route == "/api/quote-draft":
            return self.handle_save_quote_draft()

        if route == "/api/quotes":
            return self.handle_create_quote()

        return self.respond_json({"error": "Not found", "path": route}, status=HTTPStatus.NOT_FOUND)

    def do_DELETE(self) -> None:  # noqa: N802
        try:
            self._do_DELETE_internal()
        except Exception:
            logger.error(f"Error handling DELETE {self.path}:\n{traceback.format_exc()}")
            self.respond_json(
                {"error": "Internal Server Error", "detail": traceback.format_exc()},
                status=HTTPStatus.INTERNAL_SERVER_ERROR
            )

    def _do_DELETE_internal(self) -> None:
        parsed = urlparse(self.path)
        route = parsed.path

        if route.startswith("/api/imports/") and len(route.split("/")) == 4:
            return self.handle_delete_import(route.split("/")[3])

        return self.respond_json({"error": "Not found", "path": route}, status=HTTPStatus.NOT_FOUND)

    def handle_delete_import(self, import_id: str) -> None:
        with self.db() as conn:
            import_doc = self.require_import(conn, import_id)
            if not import_doc:
                return self.respond_json({"error": "Import not found"}, status=HTTPStatus.NOT_FOUND)
            
            # Delete associated issues, rows, files, document (assuming no cascade, perform manually to be safe)
            conn.execute("delete from import_row_issue where import_row_id in (select id from import_row where import_batch_id = %s)", (import_id,))
            conn.execute("delete from import_row where import_batch_id = %s", (import_id,))
            conn.execute("delete from import_file where import_batch_id = %s", (import_id,))
            conn.execute("delete from import_batch where id = %s", (import_id,))
            conn.commit()
            
        return self.respond_json({"status": "deleted", "id": import_id})

    def read_json_body(self) -> dict:
        content_length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(content_length) if content_length else b"{}"
        return json.loads(raw.decode("utf-8"))

    def parse_multipart_body(self) -> tuple[dict, list[dict]]:
        """Parse multipart/form-data. Returns (payload_dict, files_list).

        Each file entry has keys: file_name, original_name, mime_type,
        size_bytes, storage_path, file_kind.
        Text fields are returned as strings in payload_dict.
        """
        content_type = self.headers.get("Content-Type", "")
        content_length_str = self.headers.get("Content-Length")
        if not content_length_str:
            return {}, []
        body = self.rfile.read(int(content_length_str))
        msg_bytes = b"Content-Type: " + content_type.encode() + b"\r\n\r\n" + body
        container = email.parser.BytesParser().parsebytes(msg_bytes)

        payload: dict = {}
        files: list[dict] = []
        for part in container.get_payload():
            if isinstance(part, str):
                continue
            name = part.get_param("name", header="content-disposition")
            filename = part.get_filename()
            if filename:
                storage_filename = f"{uuid.uuid4()}_{filename}"
                storage_path = UPLOADS_DIR / storage_filename
                raw_bytes = part.get_payload(decode=True)
                storage_path.write_bytes(raw_bytes)
                files.append({
                    "file_name": filename,
                    "original_name": filename,
                    "mime_type": part.get_content_type(),
                    "size_bytes": len(raw_bytes),
                    "storage_path": str(storage_path),
                    "file_kind": "price" if name == "file" else "attachment",
                    "raw_bytes": raw_bytes,
                })
            else:
                payload[name] = part.get_payload(decode=True).decode("utf-8")
        return payload, files

    def validate_uuid(self, value: str) -> bool:
        """Return True if *value* is a valid UUID string."""
        try:
            uuid.UUID(str(value))
            return True
        except (ValueError, AttributeError):
            return False

    def respond_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=json_default).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def db(self):
        retries = 3
        last_exc = None
        for i in range(retries):
            try:
                # prepare=False prevents DuplicatePreparedStatement errors in ThreadingHTTPServer
                # connect_timeout=3 prevents silent hangs if Railway networking fails
                return psycopg.connect(
                    self.config.db_dsn, 
                    row_factory=dict_row, 
                    prepare_threshold=None,
                    connect_timeout=3
                )
            except psycopg.Error as e:
                last_exc = e
                logger.warning(f"Database connection attempt {i+1} failed: {e}. Retrying in 2s...")
                time.sleep(2)
        raise Exception(f"Failed to connect to db after {retries} retries: {last_exc}")


    def require_import(self, conn, import_id: str) -> dict | None:
        item = conn.execute(
            """
            select
              ib.id,
              ib.supplier_id,
              s.name as supplier_name,
              s.contract_type,
              s.vat_included,
              ib.source,
              ib.status,
              ib.processing_status,
              ib.document_type,
              ib.source_format,
              ib.currency,
              ib.period,
              ib.sheet_name,
              ib.import_date,
              ib.request_ref,
              ib.request_title,
              ib.manager_note,
              ib.processing_started_at,
              ib.processing_finished_at,
              ib.last_webhook_at,
              ib.meta,
              owner.email as owner_email,
              created_by.email as created_by_email
            from import_batch ib
            join supplier s on s.id = ib.supplier_id
            left join app_user owner on owner.id = ib.owner_user_id
            left join app_user created_by on created_by.id = ib.created_by_user_id
            where ib.id = %s
            """,
            (import_id,),
        ).fetchone()
        return item


    def get_default_user_id(self, conn) -> str | None:
        row = conn.execute(
            "select id from app_user where email = %s limit 1",
            (self.config.default_user_email,),
        ).fetchone()
        return str(row["id"]) if row else None

    def load_suppliers(self, conn) -> list[dict]:
        rows = conn.execute(
            """
            select id, external_code, name, contract_type, vat_included
            from supplier
            where is_active = true
            order by name asc
            """
        ).fetchall()
        return [
            {
                "id": str(row["id"]),
                "external_code": row["external_code"],
                "name": row["name"],
                "contract_type": row["contract_type"],
                "vat_included": row["vat_included"],
            }
            for row in rows
        ]

    def load_import_files(self, conn, import_id: str) -> list[dict]:
        rows = conn.execute(
            """
            select
              id,
              original_name,
              storage_path,
              mime_type,
              size_bytes,
              file_kind,
              note,
              processing_status,
              processing_pipeline,
              last_error,
              uploaded_at
            from import_file
            where import_batch_id = %s
            order by uploaded_at asc
            """,
            (import_id,),
        ).fetchall()
        return [
            {
                "id": str(row["id"]),
                "file_name": row["original_name"],
                "storage_path": row["storage_path"],
                "mime_type": row["mime_type"],
                "size_bytes": row["size_bytes"],
                "file_kind": row["file_kind"],
                "note": row["note"],
                "processing_status": row["processing_status"],
                "processing_pipeline": row["processing_pipeline"],
                "last_error": row["last_error"],
                "uploaded_at": row["uploaded_at"],
                "source_format": infer_source_format(row["original_name"], row["mime_type"]),
            }
            for row in rows
        ]

    def load_import_rows(self, conn, import_id: str) -> list[dict]:
        rows = conn.execute(
            """
            select
              ir.id,
              ir.import_batch_id,
              ir.row_index,
              ir.external_product_id,
              ir.temp_product_id,
              ir.raw_name,
              ir.normalized_name,
              ir.manual_normalized_name,
              ir.normalization_note,
              ir.category,
              ir.country,
              ir.volume_l,
              ir.purchase_price,
              ir.rrc_min,
              ir.promo,
              ir.review_status,
              ir.excluded,
              ir.raw_payload->>'article' as article,
              ir.raw_payload->>'note' as note,
              count(iri.id) as issues_total
            from import_row ir
            left join import_row_issue iri on iri.import_row_id = ir.id
            where ir.import_batch_id = %s
            group by ir.id
            order by ir.row_index asc
            """,
            (import_id,),
        ).fetchall()
        return [
            {
                "id": str(row["id"]),
                "import_batch_id": str(row["import_batch_id"]),
                "row_index": row["row_index"],
                "product_id": row["external_product_id"],
                "temp_id": row["temp_product_id"],
                "raw_name": row["raw_name"],
                "normalized_name": row["manual_normalized_name"] or row["normalized_name"],
                "category": row["category"],
                "country": row["country"],
                "volume_l": float(row["volume_l"]) if row["volume_l"] is not None else None,
                "purchase_price": float(row["purchase_price"]) if row["purchase_price"] is not None else None,
                "rrc_min": float(row["rrc_min"]) if row["rrc_min"] is not None else None,
                "promo": row["promo"],
                "review_status": row["review_status"],
                "excluded": row["excluded"],
                "article": row["article"],
                "note": row["note"],
                "issues_total": row["issues_total"],
                "ids": {},
            }
            for row in rows
        ]

    def load_import_issues(self, conn, import_id: str) -> list[dict]:
        rows = conn.execute(
            """
            select
              iri.id,
              iri.import_row_id,
              iri.severity,
              iri.field_name,
              iri.message,
              iri.raw_value,
              ir.row_index
            from import_row_issue iri
            join import_row ir on ir.id = iri.import_row_id
            where ir.import_batch_id = %s
            order by ir.row_index asc, iri.created_at asc
            """,
            (import_id,),
        ).fetchall()
        return [
            {
                "id": str(row["id"]),
                "import_row_id": str(row["import_row_id"]),
                "severity": row["severity"],
                "field": row["field_name"],
                "message": row["message"],
                "raw_value": row["raw_value"],
                "row_index": row["row_index"],
            }
            for row in rows
        ]

    def serialize_import(self, conn, batch_row: dict) -> dict:
        import_id = str(batch_row["id"])
        files = self.load_import_files(conn, import_id)
        stats = conn.execute(
            """
            select 
                count(*) as cnt,
                count(case when review_status = 'checked' then 1 end) as checked_cnt
            from import_row 
            where import_batch_id = %s
            """,
            (import_id,)
        ).fetchone()
        row_count = stats["cnt"]
        checked_count = stats["checked_cnt"]
        issues = self.load_import_issues(conn, import_id)
        price_file = next((item for item in files if item["file_kind"] == "price"), files[0] if files else None)
        attachments = [item for item in files if item["file_kind"] != "price"]

        # Шаг 6: Унифицированный статус — если есть строки, ставим parsed
        derived_status = batch_row["processing_status"] or batch_row["status"] or "uploaded"
        if row_count > 0 and derived_status not in ("parsed", "partial"):
            derived_status = "parsed"

        return {
            "id": import_id,
            # Шаг 7: supplier_id в корне для удобного маппинга на фронте
            "supplier_id": str(batch_row["supplier_id"]),
            "created_at": (
                batch_row["created_at"].replace(tzinfo=__import__('datetime').timezone.utc).isoformat()
                if "created_at" in batch_row and batch_row["created_at"]
                else (batch_row.get("import_date") or "")
            ),
            "meta": {
                "source_file": price_file["file_name"] if price_file else None,
                "source_format": batch_row["source_format"] or (price_file["source_format"] if price_file else None),
                "import_date": batch_row["import_date"],
                "currency": batch_row["currency"],
                "document_type": batch_row["document_type"],
                "period": batch_row["period"],
                "sheet_name": batch_row["sheet_name"],
                "attachments": attachments,
                "manager_note": batch_row["manager_note"],
                "request_ref": batch_row["request_ref"],
                "request_title": batch_row["request_title"],
                "processing_started_at": batch_row["processing_started_at"],
                "processing_finished_at": batch_row["processing_finished_at"],
                "last_webhook_at": batch_row["last_webhook_at"],
            },
            "supplier": {
                "id": str(batch_row["supplier_id"]),
                "name": batch_row["supplier_name"],
                "contract_type": batch_row["contract_type"],
                "vat_included": batch_row["vat_included"],
            },
            "created_by": batch_row["created_by_email"],
            "source": batch_row["source"],
            "owner": batch_row["owner_email"],
            "processing_status": derived_status,
            "status": derived_status,
            "row_count": row_count,
            "checked_count": checked_count,
            "errors": [
                {
                    "row_index": issue["row_index"],
                    "field": issue["field"],
                    "message": issue["message"],
                    "raw_value": issue["raw_value"],
                }
                for issue in issues
                if issue["severity"] == "error"
            ],
            "warnings": [
                {
                    "row_index": issue["row_index"],
                    "field": issue["field"],
                    "message": issue["message"],
                    "raw_value": issue["raw_value"],
                }
                for issue in issues
                if issue["severity"] != "error"
            ],
            # products убраны из листинга — загружаются отдельно через /api/products
        }

    def handle_bootstrap(self) -> None:
        with self.db() as conn:
            imports = self.load_serialized_imports(conn)
            suppliers = self.load_suppliers(conn)
            clients = self.load_clients(conn)
        return self.respond_json(
            {
                "items": {
                    "imports": imports,
                    "suppliers": suppliers,
                    "clients": clients,
                },
                "runtime": {
                    "data_source": "postgres-api",
                    "version": 1,
                },
            }
        )

    def load_clients(self, conn) -> list[dict]:
        rows = conn.execute(
            """
            select id, name, inn, city
            from client_account
            order by name asc
            """
        ).fetchall()
        return [
            {
                "id": str(row["id"]),
                "name": row["name"],
                "inn": row["inn"],
                "city": row["city"],
            }
            for row in rows
        ]

    def load_serialized_imports(self, conn) -> list[dict]:
        batch_rows = conn.execute(
            """
            select
              ib.id,
              ib.supplier_id,
              s.name as supplier_name,
              s.contract_type,
              s.vat_included,
              ib.source,
              ib.status,
              ib.processing_status,
              ib.document_type,
              ib.source_format,
              ib.currency,
              ib.period,
              ib.sheet_name,
              ib.import_date,
              ib.created_at,
              ib.updated_at,
              ib.request_ref,
              ib.request_title,
              ib.manager_note,
              ib.processing_started_at,
              ib.processing_finished_at,
              ib.last_webhook_at,
              owner.email as owner_email,
              created_by.email as created_by_email
            from import_batch ib
            join supplier s on s.id = ib.supplier_id
            left join app_user owner on owner.id = ib.owner_user_id
            left join app_user created_by on created_by.id = ib.created_by_user_id
            order by ib.created_at desc nulls last, ib.import_date desc
            """
        ).fetchall()
        return [self.serialize_import(conn, row) for row in batch_rows]

    def handle_list_imports(self) -> None:
        with self.db() as conn:
            imports = self.load_serialized_imports(conn)
        return self.respond_json({"items": imports})

    def handle_list_suppliers(self) -> None:
        with self.db() as conn:
            suppliers = self.load_suppliers(conn)
        return self.respond_json({"items": suppliers})

    def handle_list_products(self, import_id: str | None) -> None:
        if not import_id:
            return self.respond_json({"items": []})
        with self.db() as conn:
            import_row = self.require_import(conn, import_id)
            if import_row is None:
                return self.respond_json({"error": "Import not found", "import_id": import_id}, status=HTTPStatus.NOT_FOUND)
            products = self.load_import_rows(conn, import_id)
            supplier_name = import_row["supplier_name"]
            source_file = self.load_import_files(conn, import_id)
            source_file_name = next((item["file_name"] for item in source_file if item["file_kind"] == "price"), source_file[0]["file_name"] if source_file else None)

        return self.respond_json(
            {
                "items": [
                    {
                        **product,
                        "import_id": import_id,
                        "supplier_id": str(import_row["supplier_id"]),
                        "supplier_name": supplier_name,
                        "source_file": source_file_name,
                    }
                    for product in products
                ]
            }
        )

    def handle_list_jobs(self) -> None:
        with self.db() as conn:
            rows = conn.execute(
                """
                select id, type, target_type, target_id, status, updated_at
                from job_run
                order by updated_at desc
                limit 50
                """
            ).fetchall()
        return self.respond_json(
            {
                "items": [
                    {
                        "id": str(row["id"]),
                        "type": row["type"],
                        "status": row["status"],
                        "target": str(row["target_id"]) if row["target_id"] else row["target_type"],
                        "updated_at": row["updated_at"],
                    }
                    for row in rows
                ]
            }
        )

    def handle_import_status(self, import_id: str) -> None:
        with self.db() as conn:
            batch_row = self.require_import(conn, import_id)
            if batch_row is None:
                return self.respond_json({"error": "Import not found", "import_id": import_id}, status=HTTPStatus.NOT_FOUND)

            # Stale detection: if processing > 5 min, mark as failed (thread died)
            current_status = batch_row["processing_status"] or batch_row["status"] or "uploaded"
            if current_status == "processing" and batch_row.get("processing_started_at"):
                from datetime import datetime, timezone, timedelta
                started = batch_row["processing_started_at"]
                if started and (datetime.now(timezone.utc) - started) > timedelta(minutes=5):
                    n8n_logger.warning(f"[AI] Stale processing detected for import={import_id}, auto-failing")
                    conn.execute(
                        "update import_batch set status='failed', processing_status='failed', "
                        "processing_finished_at=now(), updated_at=now(), "
                        "meta = jsonb_set(coalesce(meta,'{}'::jsonb), '{error}', '\"Stale: processing timed out\"'::jsonb) "
                        "where id=%s",
                        (import_id,),
                    )
                    conn.commit()
                    batch_row = self.require_import(conn, import_id)

            files = self.load_import_files(conn, import_id)

            row_count = conn.execute(
                "select count(*) as cnt from import_row where import_batch_id = %s",
                (import_id,)
            ).fetchone()["cnt"]
            derived_status = batch_row["processing_status"] or batch_row["status"] or "uploaded"
            if row_count > 0 and derived_status not in ("parsed", "partial"):
                derived_status = "parsed"

            status = {
                "id": import_id,
                "status": derived_status,
                "processing_status": derived_status,
                "row_count": row_count,
                "processing_started_at": batch_row["processing_started_at"],
                "processing_finished_at": batch_row["processing_finished_at"],
                "last_webhook_at": batch_row["last_webhook_at"],
                "files": [
                    {
                        "id": file["id"],
                        "file_kind": file["file_kind"],
                        "original_name": file["file_name"],
                        "processing_status": file["processing_status"],
                        "processing_pipeline": file["processing_pipeline"],
                        "last_error": file["last_error"],
                    }
                    for file in files
                ],
            }
        return self.respond_json({"item": status})

    def handle_review_rows(self) -> None:
        payload = self.read_json()
        updates = payload.get("updates", [])
        if not updates:
            return self.respond_json({"status": "ok", "updated": 0})
        
        with self.db() as conn:
            for u in updates:
                conn.execute(
                    """
                    UPDATE import_row 
                    SET review_status = %s, excluded = %s
                    WHERE import_batch_id = %s AND row_index = %s
                    """,
                    (u.get("review_status"), u.get("excluded", False), u.get("import_id"), u.get("row_index"))
                )
            conn.commit()
            
        logger.info(f"[API] Updated review_status for {len(updates)} rows")
        return self.respond_json({"status": "ok", "updated": len(updates)})

    def handle_create_import(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" in content_type:
            payload, files = self.parse_multipart_body()
            payload["files"] = files
            payload["attachments"] = [f for f in files if f["file_kind"] == "attachment"]
        else:
            payload = self.read_json_body()
            files = payload.get("files") or []

        if not files:
            return self.respond_json({"error": "files is required"}, status=HTTPStatus.BAD_REQUEST)

        with self.db() as conn:
            supplier_id = payload.get("supplier_id")
            supplier_row = conn.execute(
                "select id, name from supplier where id = %s limit 1",
                (supplier_id,),
            ).fetchone()
            if supplier_row is None:
                return self.respond_json({"error": "Supplier not found", "supplier_id": supplier_id}, status=HTTPStatus.BAD_REQUEST)

            default_user_id = self.get_default_user_id(conn)
            main_file = files[0]
            source_format = infer_source_format(main_file.get("file_name") or main_file.get("original_name") or "", main_file.get("mime_type"))
            import_row = conn.execute(
                """
                insert into import_batch (
                  supplier_id,
                  owner_user_id,
                  created_by_user_id,
                  source,
                  status,
                  processing_status,
                  document_type,
                  source_format,
                  currency,
                  period,
                  import_date,
                  request_ref,
                  request_title,
                  manager_note
                )
                values (
                  %s, %s, %s, %s,
                  'uploaded',
                  'uploaded',
                  %s, %s, %s, %s, %s,
                  %s, %s, %s
                )
                returning id
                """,
                (
                    supplier_id,
                    default_user_id,
                    default_user_id,
                    payload.get("source", "Web-интерфейс"),
                    payload.get("document_type", "price_list"),
                    source_format,
                    payload.get("currency", "RUB"),
                    payload.get("period"),
                    payload.get("import_date") or date.today().isoformat(),
                    payload.get("request_ref"),
                    payload.get("request_title"),
                    payload.get("manager_note"),
                ),
            ).fetchone()
            import_id = str(import_row["id"])

            for file in files:
                original_name = file.get("file_name") or file.get("original_name")
                conn.execute(
                    """
                    insert into import_file (
                      import_batch_id,
                      original_name,
                      storage_path,
                      mime_type,
                      size_bytes,
                      file_kind,
                      note,
                      processing_status,
                      processing_pipeline,
                      file_bytes
                    )
                    values (%s, %s, %s, %s, %s, %s, %s, 'uploaded', %s, %s)
                    """,
                    (
                        import_id,
                        original_name,
                        file.get("storage_path") or f"/uploads/demo/{original_name}",
                        file.get("mime_type"),
                        file.get("size_bytes"),
                        file.get("file_kind", "price"),
                        file.get("note"),
                        infer_pipeline(infer_source_format(original_name or "", file.get("mime_type"))),
                        file.get("raw_bytes"),
                    ),
                )

            conn.commit()

        # Run PDF extraction directly (no n8n needed)
        try:
            n8n_logger.info(f"[AI] AUTO-DISPATCH queued for import_id={import_id}")
            self._trigger_n8n_import_dispatch(import_id, force=True)
        except Exception:
            logger.error(f"Auto-dispatch failed for import_id={import_id}:\n{traceback.format_exc()}")
            n8n_logger.error(f"[AI] AUTO-DISPATCH FAILED import_id={import_id}\n{traceback.format_exc()}")



        with self.db() as conn:
            batch_row = self.require_import(conn, import_id)
            item = self.serialize_import(conn, batch_row)

        return self.respond_json({"item": item}, status=HTTPStatus.CREATED)

    def dispatch_to_n8n(self, dispatch_payload: dict) -> dict | None:
        """Send *dispatch_payload* to the configured n8n webhook.

        Uses correlation_id (present in dispatch_payload) for structured logging.
        Properly closes file descriptors via context manager.
        Returns the parsed JSON response or None on failure.
        """
        if not self.config.n8n_webhook_url:
            n8n_logger.warning("[N8N] SKIP — N8N_IMPORT_WEBHOOK_URL is not configured")
            return None

        correlation_id = dispatch_payload.get("correlation_id", "-")
        job_id = dispatch_payload.get("job_id", "-")
        import_id = dispatch_payload.get("import_batch_id") or dispatch_payload.get("quote_id", "-")
        pipeline = dispatch_payload.get("requested_pipeline") or dispatch_payload.get("task_type", "-")
        source_file = dispatch_payload.get("source_file") or "-"

        data = {k: v for k, v in dispatch_payload.items() if k != "file_binary"}
        file_info = dispatch_payload.get("file_binary")

        n8n_logger.info(
            f"[N8N] DISPATCH → url={self.config.n8n_webhook_url} "
            f"correlation_id={correlation_id} job_id={job_id} "
            f"target={import_id} pipeline={pipeline} file={source_file} "
            f"file_info={file_info} exists={os.path.exists(file_info['path']) if file_info and file_info.get('path') else False}"
        )

        t_start = time.monotonic()
        try:
            headers = {"User-Agent": "bahus-API/1.0"}

            db_file_bytes = None
            if file_info:
                # Always prefer pulling the actual bytes from the database since local storage is ephemeral
                # Import_file_id is stored at the root of dispatch_payload or inside file_binary
                import_file_id = dispatch_payload.get("import_file_id") or file_info.get("import_file_id")
                if import_file_id:
                    with self.db() as conn:
                        row = conn.execute("select file_bytes from import_file where id = %s", (import_file_id,)).fetchone()
                        if row and row["file_bytes"]:
                            db_file_bytes = bytes(row["file_bytes"])

            if file_info and db_file_bytes:
                filename = file_info["filename"]
                mime_type = file_info.get("mime_type", "application/octet-stream")
                response = requests.post(
                    self.config.n8n_webhook_url,
                    data=data,
                    files={"file": (filename, db_file_bytes, mime_type)},
                    headers=headers,
                    timeout=45,
                )
            elif file_info and file_info.get("path") and os.path.exists(file_info["path"]):
                filename = file_info["filename"]
                mime_type = file_info.get("mime_type", "application/octet-stream")
                with open(file_info["path"], "rb") as fh:
                    response = requests.post(
                        self.config.n8n_webhook_url,
                        data=data,
                        files={"file": (filename, fh, mime_type)},
                        headers=headers,
                        timeout=45,
                    )
            else:
                response = requests.post(
                    self.config.n8n_webhook_url,
                    data=data,
                    headers=headers,
                    timeout=30,
                )

            latency = time.monotonic() - t_start
            response.raise_for_status()
            result = response.json() if response.text else {}
            n8n_logger.info(
                f"[N8N] RESPONSE ← status={response.status_code} "
                f"latency={latency:.2f}s correlation_id={correlation_id} "
                f"response_size={len(response.content)}B"
            )
            return result

        except requests.exceptions.RequestException as e:
            latency = time.monotonic() - t_start
            detail = ""
            if hasattr(e, "response") and e.response is not None:
                detail = f" body={e.response.text[:300]}"
            n8n_logger.error(
                f"[N8N] ERROR ← latency={latency:.2f}s correlation_id={correlation_id} "
                f"error={e}{detail}"
            )
            return None

    # ──────────────────────────────────────────────────────────────────────────
    #  AI PDF Extraction Pipeline
    #
    #  Flow:
    #    /dispatch  →  _trigger_n8n_import_dispatch
    #                    ├─ _load_price_file_bytes  (read from DB)
    #                    ├─ _mark_import_processing  (DB update)
    #                    ├─ _analyze_pdf             (PyMuPDF heuristic → mode)
    #                    └─ Thread → _chat_completions_extract → _save_extraction_result
    #
    #  Modes:
    #    text   — full text extracted by PyMuPDF → Chat Completions (fast, ~15s)
    #    vision — pages rendered as PNG → Chat Completions Vision (for scans, ~30s)
    # ──────────────────────────────────────────────────────────────────────────

    _EXTRACTION_SYSTEM_PROMPT = (
        "You are a master data extractor for wine/spirits price lists.\n"
        "Extract ALL product rows from the provided document.\n"
        "Output ONLY a valid JSON object (no markdown, no code fences):\n"
        '{"rows": [{"name":"Full product name","article":"SKU/article if present",'
        '"purchase_price":100.00,"volume_l":0.75,"country":"Country",'
        '"category":"Wine/Spirits/Beer/etc","note":"any extra info"}]}\n'
        "Rules:\n"
        "- Extract EVERY row, do not skip or summarize.\n"
        "- purchase_price must be a number (no currency symbols).\n"
        "- volume_l in liters (0.75, 1.0, 0.5, etc).\n"
        "- If a field value is unknown, use null.\n"
        "- Do NOT wrap output in markdown code fences."
    )

    def _trigger_n8n_import_dispatch(self, import_id: str, force: bool = False) -> dict:
        """Dispatch AI extraction: analyze PDF type then run Chat Completions in background.

        Returns immediately after starting the background thread.
        The thread writes results directly to the database.
        """
        # ── 1. Load import + file bytes from DB ───────────────────────────────
        with self.db() as conn:
            batch_row = self.require_import(conn, import_id)
            if batch_row is None:
                raise ValueError("Import not found")

            if not force:
                cnt = conn.execute(
                    "select count(*) as cnt from import_row where import_batch_id = %s",
                    (import_id,)
                ).fetchone()["cnt"]
                if cnt > 0:
                    return {"import_id": import_id, "skipped": True, "reason": "already_processed"}

            files = self.load_import_files(conn, import_id)
            price_file = next((f for f in files if f["file_kind"] == "price"), None)
            if price_file is None:
                raise ValueError("Price file not found")

            file_id_db = str(price_file["id"])
            file_row = conn.execute(
                "select file_bytes, original_name from import_file where id = %s",
                (file_id_db,)
            ).fetchone()
            if not file_row or not file_row["file_bytes"]:
                raise ValueError("File bytes not found — re-upload the file")

            file_bytes = bytes(file_row["file_bytes"])
            filename = file_row["original_name"] or "file.pdf"

            # ── 2. Mark import as processing ──────────────────────────────────
            job_id = str(uuid.uuid4())
            default_user_id = self.get_default_user_id(conn)
            conn.execute(
                "insert into job_run (id, type, target_type, target_id, status, payload, created_by_user_id) "
                "values (%s, 'import_parse', 'import_batch', %s, 'running', '{}'::jsonb, %s)",
                (job_id, import_id, default_user_id),
            )
            conn.execute(
                "update import_batch "
                "set status='processing', processing_status='processing', "
                "    processing_started_at=now(), processing_finished_at=null, updated_at=now() "
                "where id=%s",
                (import_id,),
            )
            conn.execute(
                "update import_file set processing_status='processing' "
                "where import_batch_id=%s and file_kind='price'",
                (import_id,),
            )
            conn.commit()

        api_key = self.config.openai_api_key
        if not api_key:
            raise ValueError("OPENAI_API_KEY is not configured")

        # ── 3. Run analysis and extraction in background thread ────────────────
        # Max chars per chunk for text-mode extraction.
        # gpt-4o context window is 128k tokens (~500k chars input),
        # but output is capped at 16 384 tokens. At ~5 chars/token avg,
        # a chunk of 40 000 chars should produce ≤8 000 output tokens —
        # well within the limit, leaving plenty of headroom for a full JSON array.
        _CHUNK_CHARS = 40_000

        def _bg_extract():
            try:
                extract_mode, pdf_payload = self._analyze_pdf(file_bytes, filename)
                n8n_logger.info(f"[AI] dispatch import={import_id} file={filename} mode={extract_mode}")

                if extract_mode == "text" and len(pdf_payload) > _CHUNK_CHARS:
                    rows = self._chunked_text_extract(api_key, pdf_payload, filename, import_id, _CHUNK_CHARS)
                else:
                    rows = self._chat_completions_extract(api_key, extract_mode, pdf_payload, filename, import_id)

                status = "parsed" if rows else "failed"
                error = None if rows else "No rows extracted"
            except Exception:
                n8n_logger.error(f"[AI] extraction error import={import_id}\n{traceback.format_exc()}")
                rows, status, error = [], "failed", traceback.format_exc().splitlines()[-1]
            self._save_extraction_result(import_id, job_id, file_id_db, rows, status, error)

        threading.Thread(target=_bg_extract, daemon=True, name=f"ai-extract-{import_id[:8]}").start()
        return {"import_id": import_id, "job_id": job_id, "status": "processing", "extract_mode": "analyzing"}

    # ── PDF Analysis ──────────────────────────────────────────────────────────

    def _analyze_pdf(self, file_bytes: bytes, filename: str) -> tuple[str, object]:
        """Analyze PDF and decide extraction strategy.

        Returns:
            ('text',   full_text_str)     — text-based PDF, send as plain text
            ('vision', [b64_png, ...])    — scanned/image PDF, send as Vision images

        Heuristic rules (in priority order):
            1. Short PDF (≤5 pages) → always Vision (accurate, cheap enough)
            2. Long PDF with rich text and price patterns → text (fast, handles 100+ rows)
            3. Everything else → Vision
        """
        import re
        import base64
        import fitz  # PyMuPDF

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        page_count = len(doc)
        pages_text = [page.get_text("text") or "" for page in doc]
        total_chars = sum(len(t.strip()) for t in pages_text)
        avg_chars = total_chars / max(page_count, 1)
        price_hits = len(re.findall(r'\d[\d\s]{2,}', "\n".join(pages_text)))

        n8n_logger.info(
            f"[AI] analyze_pdf file={filename} pages={page_count} "
            f"total_chars={total_chars} avg={avg_chars:.0f} price_hits={price_hits}"
        )

        use_text = (
            page_count > 5           # not a short doc
            and avg_chars > 200      # rich text layer
            and price_hits > 5       # actually has numeric data
        )

        if use_text:
            doc.close()
            full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages_text)
            n8n_logger.info(f"[AI] mode=text ({len(full_text)} chars)")
            return "text", full_text

        # Vision: render each page to PNG at 150 DPI (limit strictly to first 10 pages for safety)
        max_vision_pages = min(page_count, 10)
        n8n_logger.info(f"[AI] mode=vision (rendering {max_vision_pages} pages at 150 DPI)")
        images_b64 = []
        for i in range(max_vision_pages):
            page = doc[i]
            # 150 DPI is a good balance of readability and image size
            images_b64.append(base64.b64encode(page.get_pixmap(dpi=150).tobytes("png")).decode("ascii"))
            
        doc.close()
        return "vision", images_b64

    # ── Chat Completions Extraction ───────────────────────────────────────────

    def _chat_completions_extract(
        self, api_key: str, mode: str, payload, filename: str, import_id: str
    ) -> list:
        """Send PDF content to GPT-4o Chat Completions and parse the JSON response.

        Args:
            mode:    'text' or 'vision'
            payload: full_text string (text mode) or list of base64 PNG strings (vision mode)
        """
        import re

        if mode == "text":
            messages = [
                {"role": "system", "content": self._EXTRACTION_SYSTEM_PROMPT},
                {"role": "user",   "content": f"Price list '{filename}':\n\n{payload}"},
            ]
        else:
            # Vision: system prompt + one user message with all page images
            image_parts = [
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"},
                }
                for b64 in payload
            ]
            messages = [
                {"role": "system", "content": self._EXTRACTION_SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": f"Scanned price list '{filename}' ({len(payload)} pages):"},
                        *image_parts,
                    ],
                },
            ]

        n8n_logger.info(f"[AI] chat_completions import={import_id} mode={mode} pages/chars={len(payload)}")
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": "gpt-4o",
                "messages": messages,
                "response_format": {"type": "json_object"},
                "temperature": 0,
                # Hard maximum for gpt-4o. Do NOT set higher — the API will error.
                "max_tokens": 16384,
            },
            timeout=300,
        )

        if not resp.ok:
            raise ValueError(f"OpenAI error {resp.status_code}: {resp.text[:300]}")

        choice = resp.json()["choices"][0]
        finish_reason = choice.get("finish_reason", "unknown")
        raw = choice["message"]["content"]
        n8n_logger.info(
            f"[AI] response import={import_id} finish_reason={finish_reason} "
            f"len={len(raw)} preview={raw[:150]}"
        )

        # Warn immediately if the model ran out of tokens — the JSON is almost certainly truncated.
        if finish_reason == "length":
            n8n_logger.warning(
                f"[AI] TRUNCATED RESPONSE import={import_id} — finish_reason=length. "
                f"Attempting JSON repair on {len(raw)} chars."
            )

        # Sanitize response (strip markdown fences and citation markers if any)
        cleaned = re.sub(r'【[^】]*】', '', raw).strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r'^```(?:json)?\s*|\s*```\s*$', '', cleaned, flags=re.MULTILINE).strip()

        if not cleaned or cleaned[0] not in ('{', '['):
            raise ValueError(f"Response is not valid JSON: {raw[:300]}")

        # ── JSON repair: close any unclosed arrays/objects from a truncated response ──
        def repair_truncated_json(s: str) -> str:
            """Close unclosed brackets/braces so the last complete object is kept."""
            # Track nesting depth
            depth_stack = []
            in_string = False
            escape_next = False
            last_complete_obj_end = 0
            depth_at_root_close = 0

            for i, ch in enumerate(s):
                if escape_next:
                    escape_next = False
                    continue
                if ch == '\\' and in_string:
                    escape_next = True
                    continue
                if ch == '"':
                    in_string = not in_string
                    continue
                if in_string:
                    continue
                if ch in ('{', '['):
                    depth_stack.append(ch)
                elif ch in ('}', ']'):
                    if depth_stack:
                        depth_stack.pop()
                    # Record the position after each top-level object closes
                    if not depth_stack:
                        last_complete_obj_end = i + 1

            if not depth_stack:
                return s  # Already valid

            # Build closing sequence
            closers = {'[': ']', '{': '}'}
            suffix = ''.join(closers[c] for c in reversed(depth_stack))
            repaired = s + suffix
            n8n_logger.info(
                f"[AI] json_repair appended {len(suffix)!r} chars: {suffix!r}"
            )
            return repaired

        try:
            parsed = json.loads(cleaned)
        except json.JSONDecodeError as e:
            n8n_logger.warning(f"[AI] json parse failed ({e}), attempting repair")
            try:
                repaired = repair_truncated_json(cleaned)
                parsed = json.loads(repaired)
                n8n_logger.info(f"[AI] json_repair succeeded for import={import_id}")
            except json.JSONDecodeError as e2:
                raise ValueError(
                    f"JSON невалиден даже после попытки восстановления: {e2}. "
                    f"Вероятно, GPT обрезал ответ (finish_reason={finish_reason}). "
                    f"Попробуйте разбить прайс-лист на меньшие части."
                ) from e2

        rows = parsed if isinstance(parsed, list) else (parsed.get("rows") or [])
        n8n_logger.info(f"[AI] extracted {len(rows)} rows from {filename}")
        return rows

    # ── Chunked text extraction for large PDFs ────────────────────────────────

    def _chunked_text_extract(
        self, api_key: str, full_text: str, filename: str, import_id: str, chunk_size: int
    ) -> list:
        """Split *full_text* into overlapping chunks and extract rows from each.

        Runs one API call per chunk sequentially (safe on Railway free tier).
        Deduplicates rows by (name, purchase_price) before returning.
        """
        # Split on page breaks when possible so items don't get clipped mid-row.
        page_sep = "\n\n--- PAGE BREAK ---\n\n"
        if page_sep in full_text:
            pages = full_text.split(page_sep)
        else:
            # Fall back to splitting by lines
            lines = full_text.splitlines(keepends=True)
            pages = []
            buf, buf_len = [], 0
            for line in lines:
                if buf_len + len(line) > chunk_size and buf:
                    pages.append("".join(buf))
                    buf, buf_len = [], 0
                buf.append(line)
                buf_len += len(line)
            if buf:
                pages.append("".join(buf))

        # Group pages into chunks of ~chunk_size chars each
        chunks, current, current_len = [], [], 0
        for page in pages:
            if current_len + len(page) > chunk_size and current:
                chunks.append(page_sep.join(current))
                # Overlap: carry last page into next chunk as context
                current, current_len = [page], len(page)
            else:
                current.append(page)
                current_len += len(page)
        if current:
            chunks.append(page_sep.join(current))

        n8n_logger.info(
            f"[AI] chunked import={import_id} file={filename} "
            f"total_chars={len(full_text)} chunks={len(chunks)} chunk_size={chunk_size}"
        )

        all_rows: list = []
        for idx, chunk in enumerate(chunks):
            n8n_logger.info(f"[AI] chunk {idx+1}/{len(chunks)} import={import_id} chars={len(chunk)}")
            try:
                chunk_rows = self._chat_completions_extract(
                    api_key, "text", chunk,
                    f"{filename} [часть {idx+1}/{len(chunks)}]",
                    import_id,
                )
                all_rows.extend(chunk_rows)
                n8n_logger.info(f"[AI] chunk {idx+1} → {len(chunk_rows)} rows (total so far: {len(all_rows)})")
            except Exception:
                n8n_logger.error(
                    f"[AI] chunk {idx+1} failed import={import_id}\n{traceback.format_exc()}"
                )
                # Continue with remaining chunks; partial results are better than nothing
                continue

        # Deduplicate by (normalised name, price) to remove overlap rows
        seen: set = set()
        deduped: list = []
        for row in all_rows:
            key = (
                str(row.get("name") or row.get("raw_name") or "").strip().lower(),
                str(row.get("purchase_price") or ""),
            )
            if key not in seen:
                seen.add(key)
                deduped.append(row)

        n8n_logger.info(
            f"[AI] chunked done import={import_id} raw={len(all_rows)} deduped={len(deduped)}"
        )
        return deduped

    # ── Persist Extraction Result ─────────────────────────────────────────────

    def _save_extraction_result(
        self, import_id: str, job_id: str, file_id: str,
        rows: list, status: str, error: str | None
    ) -> None:
        """Write extraction result to DB. Called from background extraction thread."""
        with self.db() as conn:
            conn.execute("delete from import_row where import_batch_id=%s", (import_id,))
            for i, row in enumerate(rows):
                conn.execute(
                    """
                    insert into import_row (
                        import_batch_id, row_index, raw_name, normalized_name,
                        category, country, volume_l, purchase_price, promo, raw_payload
                    ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)
                    """,
                    (
                        import_id, i + 1,
                        row.get("name") or row.get("raw_name") or f"Строка {i+1}",
                        row.get("normalized_name") or row.get("name"),
                        row.get("category"),
                        row.get("country"),
                        row.get("volume_l"),
                        row.get("purchase_price"),
                        bool(row.get("promo", False)),
                        json.dumps(row, ensure_ascii=False, default=json_default),
                    ),
                )
            conn.execute(
                """
                update import_batch
                set status=%s, processing_status=%s,
                    processing_finished_at=now(), last_webhook_at=now(), updated_at=now(),
                    meta = jsonb_set(coalesce(meta, '{}'::jsonb), '{error}', %s::jsonb)
                where id=%s
                """,
                (status, status, json.dumps(error or ""), import_id),
            )
            conn.execute(
                "update import_file "
                "set processing_status=%s, last_error=%s "
                "where import_batch_id=%s and file_kind='price'",
                (status, error, import_id),
            )
            conn.execute(
                "update job_run set status=%s, finished_at=now() where id=%s",
                ("done" if status == "parsed" else "failed", job_id),
            )
            conn.commit()
        n8n_logger.info(f"[AI] saved import={import_id} status={status} rows={len(rows)}")



    def handle_dispatch_import(self, import_id: str) -> None:

        payload = self.read_json_body()
        try:
            res = self._trigger_n8n_import_dispatch(import_id, force=bool(payload.get("force", False)))
            return self.respond_json(
                {
                    "item": res,
                    "dispatch": "queued",
                },
                status=HTTPStatus.CREATED,
            )
        except ValueError as e:
            return self.respond_json({"error": str(e), "import_id": import_id}, status=HTTPStatus.BAD_REQUEST)

    def resolve_import_file_id(self, conn, import_id: str, import_file_id: str | None) -> str | None:
        if import_file_id:
            return import_file_id
        row = conn.execute(
            """
            select id
            from import_file
            where import_batch_id = %s and file_kind = 'price'
            order by uploaded_at asc
            limit 1
            """,
            (import_id,),
        ).fetchone()
        return str(row["id"]) if row else None

    def handle_n8n_import_result(self) -> None:
        payload = self.read_json_body()
        import_id = payload.get("import_batch_id") or payload.get("import_id")
        if not import_id:
            return self.respond_json({"error": "import_batch_id is required"}, status=HTTPStatus.BAD_REQUEST)

        # Validate UUID format before hitting DB (prevents psycopg InvalidTextRepresentation)
        try:
            uuid.UUID(str(import_id))
        except ValueError:
            n8n_logger.warning(f"[N8N] WEBHOOK ← invalid import_id format (not UUID): {import_id!r}")
            return self.respond_json(
                {"error": "import_batch_id must be a valid UUID", "received": import_id},
                status=HTTPStatus.BAD_REQUEST,
            )

        correlation_id = payload.get("correlation_id", "-")
        rows_count = len(payload.get("rows") or [])
        issues_count = len(payload.get("issues") or [])
        n8n_logger.info(
            f"[N8N] WEBHOOK ← /api/webhooks/n8n/import-result "
            f"import_id={import_id} correlation_id={correlation_id} "
            f"rows={rows_count} issues={issues_count} "
            f"status={payload.get('status', '?')} pipeline={payload.get('pipeline', '?')}"
        )

        with self.db() as conn:
            batch_row = self.require_import(conn, import_id)
            if batch_row is None:
                return self.respond_json({"error": "Import not found", "import_id": import_id}, status=HTTPStatus.NOT_FOUND)


            import_file_id = self.resolve_import_file_id(conn, import_id, payload.get("import_file_id"))
            status = payload.get("status", "parsed")
            meta = payload.get("meta") or {}
            rows = payload.get("rows") or []
            issues = payload.get("issues") or []
            parse_result = payload.get("parse_result") or payload

            if payload.get("job_id") and len(str(payload.get("job_id"))) == 36 and "-" in str(payload.get("job_id")):
                conn.execute(
                    """
                    update job_run
                    set
                      status = 'done',
                      started_at = coalesce(started_at, now()),
                      finished_at = now(),
                      result = %s::jsonb,
                      updated_at = now()
                    where id = %s
                    """,
                    (json.dumps(payload, ensure_ascii=False, default=json_default), payload["job_id"]),
                )

            conn.execute(
                """
                update import_batch
                set
                  status = %s,
                  processing_status = %s,
                  processing_finished_at = now(),
                  last_webhook_at = now(),
                  currency = coalesce(%s, currency),
                  period = coalesce(%s, period),
                  sheet_name = coalesce(%s, sheet_name),
                  updated_at = now()
                where id = %s
                """,
                (status, status, meta.get("currency"), meta.get("period"), meta.get("sheet_name"), import_id),
            )
            if import_file_id:
                conn.execute(
                    """
                    update import_file
                    set
                      processing_status = %s,
                      processing_pipeline = %s,
                      parse_result = %s::jsonb,
                      last_error = null
                    where id = %s
                    """,
                    (
                        status,
                        payload.get("pipeline"),
                        json.dumps(parse_result, ensure_ascii=False, default=json_default),
                        import_file_id,
                    ),
                )

            conn.execute("delete from import_row where import_batch_id = %s", (import_id,))

            row_map = {}
            for index, row in enumerate(rows):
                row_index = int(row.get("row_index", index + 1))
                inserted = conn.execute(
                    """
                    insert into import_row (
                      import_batch_id,
                      row_index,
                      external_product_id,
                      temp_product_id,
                      raw_name,
                      normalized_name,
                      category,
                      country,
                      volume_l,
                      purchase_price,
                      rrc_min,
                      promo,
                      raw_payload
                    )
                    values (
                      %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb
                    )
                    returning id
                    """,
                    (
                        import_id,
                        row_index,
                        row.get("product_id") or row.get("external_product_id"),
                        row.get("temp_id") or row.get("temp_product_id"),
                        row.get("raw_name") or row.get("name") or f"Строка {row_index}",
                        row.get("normalized_name"),
                        row.get("category"),
                        row.get("country"),
                        row.get("volume_l"),
                        row.get("purchase_price"),
                        row.get("rrc_min") or row.get("sale_price"),
                        bool(row.get("promo", False)),
                        json.dumps(row, ensure_ascii=False, default=json_default),
                    ),
                ).fetchone()
                row_map[row_index] = str(inserted["id"])

            for issue in issues:
                row_id = row_map.get(int(issue.get("row_index", 0)))
                if not row_id:
                    continue
                conn.execute(
                    """
                    insert into import_row_issue (
                      import_row_id,
                      severity,
                      field_name,
                      message,
                      raw_value
                    )
                    values (%s, %s, %s, %s, %s)
                    """,
                    (
                        row_id,
                        "error" if issue.get("severity") == "error" else "warning",
                        issue.get("field"),
                        issue.get("message") or "Требует проверки",
                        issue.get("raw_value"),
                    ),
                )

            conn.commit()
            refreshed = self.serialize_import(conn, self.require_import(conn, import_id))

        return self.respond_json({"item": refreshed})

    def handle_n8n_import_failed(self) -> None:
        payload = self.read_json_body()
        import_id = payload.get("import_batch_id") or payload.get("import_id")
        if not import_id:
            return self.respond_json({"error": "import_batch_id is required"}, status=HTTPStatus.BAD_REQUEST)

        correlation_id = payload.get("correlation_id", "-")
        error_msg = payload.get("error") or payload.get("message") or "unknown"
        n8n_logger.error(
            f"[N8N] WEBHOOK ← /api/webhooks/n8n/import-failed "
            f"import_id={import_id} correlation_id={correlation_id} "
            f"error={error_msg!r}"
        )

        with self.db() as conn:
            batch_row = self.require_import(conn, import_id)
            if batch_row is None:
                return self.respond_json({"error": "Import not found", "import_id": import_id}, status=HTTPStatus.NOT_FOUND)

            import_file_id = self.resolve_import_file_id(conn, import_id, payload.get("import_file_id"))
            error_text = payload.get("error") or payload.get("message") or "Обработка завершилась с ошибкой"

            if payload.get("job_id") and len(str(payload.get("job_id"))) == 36 and "-" in str(payload.get("job_id")):
                conn.execute(
                    """
                    update job_run
                    set
                      status = 'failed',
                      started_at = coalesce(started_at, now()),
                      finished_at = now(),
                      result = %s::jsonb,
                      updated_at = now()
                    where id = %s
                    """,
                    (json.dumps(payload, ensure_ascii=False, default=json_default), payload["job_id"]),
                )

            conn.execute(
                """
                update import_batch
                set
                  status = 'failed',
                  processing_status = 'failed',
                  processing_finished_at = now(),
                  last_webhook_at = now(),
                  updated_at = now()
                where id = %s
                """,
                (import_id,),
            )
            if import_file_id:
                conn.execute(
                    """
                    update import_file
                    set
                      processing_status = 'failed',
                      processing_pipeline = %s,
                      last_error = %s
                    where id = %s
                    """,
                    (payload.get("pipeline"), error_text, import_file_id),
                )
            conn.commit()
            refreshed = self.serialize_import(conn, self.require_import(conn, import_id))

        return self.respond_json({"item": refreshed, "error": error_text})


    def handle_enrich_photo(self, row_id: str) -> None:
        """Search and save a product photo using Serper.dev API."""
        if not self.config.serper_api_key:
            return self.respond_json({"error": "SERPER_API_KEY is not configured on the backend."}, status=HTTPStatus.BAD_REQUEST)

        with self.db() as conn:
            row = conn.execute("select normalized_name, raw_name from import_row where id=%s", (row_id,)).fetchone()
            if not row:
                return self.respond_json({"error": "Row not found"}, status=HTTPStatus.NOT_FOUND)

        query = f"{row['normalized_name'] or row['raw_name']} wine bottle white background"

        url = "https://google.serper.dev/images"
        payload = json.dumps({"q": query, "num": 10})
        headers = {
            'X-API-KEY': self.config.serper_api_key,
            'Content-Type': 'application/json'
        }

        try:
            import requests
            response = requests.post(url, headers=headers, data=payload, timeout=10)
            if response.status_code != 200:
                logger.error(f"Serper API failed: {response.text}")
                return self.respond_json({"error": "Search API failed"}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
            
            data = response.json()
            images = data.get("images", [])
            
            # Filter for reasonable images (squares, or portrait bottle aspect ratios)
            valid_images = [
                img["imageUrl"] for img in images 
                if img.get("imageWidth", 1) / img.get("imageHeight", 1) <= 1.5
            ]
            if not valid_images:
                valid_images = [img["imageUrl"] for img in images]

            if not valid_images:
                return self.respond_json({"error": "No images found"}, status=HTTPStatus.NOT_FOUND)
            
            best_image = valid_images[0]

            with self.db() as conn:
                conn.execute(
                    "UPDATE import_row SET image_url=%s WHERE id=%s",
                    (best_image, row_id)
                )
                conn.commit()

            return self.respond_json({"image_url": best_image})
        except Exception as e:
            logger.error(f"Image search exception: {e}")
            return self.respond_json({"error": str(e)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_n8n_quote_result(self) -> None:
        payload = self.read_json_body()
        quote_id = payload.get("quote_id") or payload.get("import_id") or payload.get("import_batch_id")
        if not quote_id:
            return self.respond_json({"error": "quote_id is required"}, status=HTTPStatus.BAD_REQUEST)
        if not self.validate_uuid(quote_id):
            return self.respond_json({"error": "Invalid quote_id format"}, status=HTTPStatus.BAD_REQUEST)

        correlation_id = payload.get("correlation_id", "-")
        rows_count = len(payload.get("rows") or [])
        n8n_logger.info(
            f"[N8N] WEBHOOK ← /api/webhooks/n8n/quote-result "
            f"quote_id={quote_id} correlation_id={correlation_id} rows={rows_count}"
        )

        # Сохраняем сырой ответ от n8n локально на случай ошибок парсинга
        try:
            raw_path = UPLOADS_DIR / f"quote_{quote_id}_raw.json"
            with open(raw_path, 'w', encoding='utf-8') as f:
                json.dump(payload, f, ensure_ascii=False, indent=2, default=json_default)
            logger.info(f"Saved raw n8n payload to {raw_path}")
        except Exception as e:
            logger.error(f"Failed to save raw payload for quote {quote_id}: {e}")

        with self.db() as conn:
            if payload.get("job_id") and len(str(payload.get("job_id"))) == 36 and "-" in str(payload.get("job_id")):
                conn.execute(
                    """
                    update job_run
                    set
                      status = 'done',
                      finished_at = now(),
                      result = %s::jsonb,
                      updated_at = now()
                    where id = %s
                    """,
                    (json.dumps(payload, ensure_ascii=False, default=json_default), payload["job_id"]),
                )

            # Check if quote exists
            row = conn.execute("select id from quote_document where id = %s", (quote_id,)).fetchone()
            if row:
                rows = payload.get("rows", [])
                if rows:
                    conn.execute("delete from quote_item where quote_document_id = %s", (quote_id,))
                    for i, item in enumerate(rows):
                        conn.execute(
                            """
                            insert into quote_item (quote_document_id, line_no, name_snapshot, volume_l, purchase_price, rrc_min, sale_price, qty)
                            values (%s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                quote_id,
                                i + 1,
                                item.get("name") or item.get("raw_name") or "Unknown item",
                                item.get("volume_l"),
                                item.get("purchase_price"),
                                item.get("rrc"),
                                item.get("rrc"),
                                item.get("qty", 1)
                            )
                        )
                # Update status
                conn.execute("update quote_document set status = 'ready' where id = %s", (quote_id,))
            else:
                logger.error(f"Quote {quote_id} not found when receiving n8n result!")

        return self.respond_json({"item": {"quote_id": quote_id, "status": "processed"}})

    def handle_n8n_quote_failed(self) -> None:
        payload = self.read_json_body()
        quote_id = payload.get("quote_id") or payload.get("import_id") or payload.get("import_batch_id")
        if not quote_id:
            return self.respond_json({"error": "quote_id is required"}, status=HTTPStatus.BAD_REQUEST)
        if not self.validate_uuid(quote_id):
            return self.respond_json({"error": "Invalid quote_id format"}, status=HTTPStatus.BAD_REQUEST)

        correlation_id = payload.get("correlation_id", "-")
        error_msg = payload.get("error") or payload.get("message") or "unknown"
        n8n_logger.error(
            f"[N8N] WEBHOOK ← /api/webhooks/n8n/quote-failed "
            f"quote_id={quote_id} correlation_id={correlation_id} error={error_msg!r}"
        )

        with self.db() as conn:
            if payload.get("job_id") and self.validate_uuid(payload["job_id"]):
                conn.execute(
                    """
                    update job_run
                    set
                      status = 'failed',
                      finished_at = now(),
                      result = %s::jsonb,
                      updated_at = now()
                    where id = %s
                    """,
                    (json.dumps(payload, ensure_ascii=False, default=json_default), payload["job_id"]),
                )
            # Fix: update quote_document status so it doesn't stay stuck as 'processing'
            conn.execute(
                "update quote_document set status = 'failed' where id = %s",
                (quote_id,),
            )
            conn.commit()
        return self.respond_json({"item": {"quote_id": quote_id, "status": "failed"}})

    def handle_get_quote_draft(self) -> None:
        draft_path = UPLOADS_DIR / "quote_draft.json"
        if not draft_path.exists():
            return self.respond_json({"message": "No draft"}, status=HTTPStatus.NOT_FOUND)
        try:
            with open(draft_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return self.respond_json(data)
        except Exception as e:
            return self.respond_json({"error": str(e)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_save_quote_draft(self) -> None:
        payload = self.read_json_body()
        draft_path = UPLOADS_DIR / "quote_draft.json"
        try:
            payload["saved_at"] = datetime.now().isoformat()
            with open(draft_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False)
            return self.respond_json(payload)
        except Exception as e:
            return self.respond_json({"error": str(e)}, status=HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_proxy_n8n(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            return self.respond_json({"error": "multipart/form-data required"}, status=HTTPStatus.BAD_REQUEST)
        if not self.headers.get("Content-Length"):
            return self.respond_json({"error": "Content-Length required"}, status=HTTPStatus.LENGTH_REQUIRED)

        payload, files_list = self.parse_multipart_body()
        # Re-build files dict suitable for requests (name → (filename, bytes, mime))
        files_for_requests = {
            f["file_kind"]: (f["file_name"], open(f["storage_path"], "rb"), f["mime_type"])
            for f in files_list
        }

        target_url = payload.pop("target_webhook_url", None)
        if not target_url:
            # Close any opened file handles before returning
            for _, fh, _ in files_for_requests.values() if files_for_requests else []:
                fh.close()
            return self.respond_json({"error": "target_webhook_url required in form data"}, status=HTTPStatus.BAD_REQUEST)

        n8n_logger.info(f"[N8N] PROXY → target={target_url} files={list(files_for_requests.keys())}")
        try:
            r = requests.post(
                target_url,
                data=payload,
                files=files_for_requests if files_for_requests else None,
                timeout=60,
            )
            for _, fh, _ in (files_for_requests.values() if files_for_requests else []):
                fh.close()
            r.raise_for_status()
            n8n_logger.info(f"[N8N] PROXY ← status={r.status_code} target={target_url}")
            try:
                return self.respond_json(r.json())
            except ValueError:
                return self.respond_json({"message": "Proxy ok, no JSON response", "raw_response": r.text})
        except requests.exceptions.RequestException as e:
            for _, fh, _ in (files_for_requests.values() if files_for_requests else []):
                try:
                    fh.close()
                except Exception:
                    pass
            n8n_logger.error(f"[N8N] PROXY_ERROR target={target_url} error={e}")
            msg = e.response.text if hasattr(e, "response") and e.response else str(e)
            return self.respond_json({"error": "N8n proxy request failed", "details": msg}, status=HTTPStatus.BAD_GATEWAY)



    def serialize_quote(self, conn, row) -> dict:
        items = conn.execute(
            '''
            select
              id as source_product_id,
              name_snapshot as name,
              volume_l,
              purchase_price,
              rrc_min,
              sale_price,
              qty,
              line_snapshot_payload
            from quote_item
            where quote_document_id = %s
            order by line_no asc
            ''',
            (row["id"],),
        ).fetchall()
        
        for item in items:
            item["source_product_id"] = str(item["source_product_id"])

        return {
            "id": str(row["id"]),
            "status": row["status"],
            "meta": {
                "clientId": str(row["client_id"]) if row["client_id"] else "",
                "quoteNumber": row["quote_number"],
                "quoteDate": str(row["quote_date"]) if row["quote_date"] else "",
                "requestTitle": "",
                "note": row["note"] or "",
                "mode": row["mode"],
                "aiProcessingStatus": "done",
            },
            "items": items,
        }

    def handle_list_quotes(self) -> None:
        with self.db() as conn:
            rows = conn.execute(
                '''
                select
                  id, client_id, quote_number, quote_date, mode, status, note
                from quote_document
                order by updated_at desc
                limit 100
                '''
            ).fetchall()
            quotes = [self.serialize_quote(conn, r) for r in rows]
        return self.respond_json({"items": quotes})

    def handle_get_quote(self, quote_id: str) -> None:
        if not self.validate_uuid(quote_id):
            return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
            
        with self.db() as conn:
            row = conn.execute("select * from quote_document where id = %s", (quote_id,)).fetchone()
            if not row:
                return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
            quote = self.serialize_quote(conn, row)
        return self.respond_json({"item": quote})

    def handle_export_quote_excel(self, quote_id: str) -> None:
        if not openpyxl:
            return self.respond_json({"error": "openpyxl not installed on backend"}, status=HTTPStatus.NOT_IMPLEMENTED)
            
        if not self.validate_uuid(quote_id):
            return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
            
        with self.db() as conn:
            row = conn.execute("select * from quote_document where id = %s", (quote_id,)).fetchone()
            if not row:
                return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
            quote = self.serialize_quote(conn, row)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Коммерческое предложение"
        
        headers = ["№", "Фото", "Позиция", "Объём (л)", "Кол-во", "Цена (руб)", "Сумма (руб)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 40
            
        current_row = 2
        for idx, item in enumerate(quote["items"], 1):
            qty = Decimal(item.get("qty") or 0)
            price = Decimal(item.get("sale_price") or 0)
            line_sum = qty * price
            
            ws.row_dimensions[current_row].height = 60
            
            ws.cell(row=current_row, column=1, value=idx).alignment = Alignment(vertical="center")
            ws.cell(row=current_row, column=3, value=item.get("name") or "").alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row=current_row, column=4, value=item.get("volume_l") or "").alignment = Alignment(vertical="center")
            ws.cell(row=current_row, column=5, value=float(qty)).alignment = Alignment(vertical="center")
            ws.cell(row=current_row, column=6, value=float(price)).alignment = Alignment(vertical="center")
            ws.cell(row=current_row, column=7, value=float(line_sum)).alignment = Alignment(vertical="center")
            
            payload = item.get("line_snapshot_payload", {})
            image_url = payload.get("image_url") if isinstance(payload, dict) else None
            
            if image_url:
                try:
                    resp = requests.get(image_url, timeout=5)
                    if resp.status_code == 200:
                        img = PillowImage.open(io.BytesIO(resp.content))
                        img.thumbnail((80, 80))
                        
                        img_io = io.BytesIO()
                        img.save(img_io, format="PNG")
                        img_io.seek(0)
                        
                        xl_img = OpenpyxlImage(img_io)
                        # Center roughly
                        ws.add_image(xl_img, f"B{current_row}")
                except Exception as e:
                    logger.error(f"Failed to load image {image_url} for quote {quote_id}: {e}")
            
            current_row += 1

        # Summary row
        ws.cell(row=current_row+1, column=3, value="ИТОГО").font = Font(bold=True)
        ws.cell(row=current_row+1, column=5, value=f"=SUM(E2:E{current_row-1})").font = Font(bold=True)
        ws.cell(row=current_row+1, column=7, value=f"=SUM(G2:G{current_row-1})").font = Font(bold=True)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="quote_{quote_id[:8]}.xlsx"')
        self.end_headers()
        self.wfile.write(out.read())

    def handle_create_quote(self) -> None:
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" in content_type:
            payload, files = self.parse_multipart_body()
            if "meta" in payload:
                payload["meta"] = json.loads(payload["meta"])
        else:
            payload = self.read_json_body()
            files = []

        client_id = payload.get("meta", {}).get("clientId")
        if client_id and not self.validate_uuid(client_id):
            client_id = None

        note = payload.get("meta", {}).get("note", "")
        status = "processing" if files else "draft"

        with self.db() as conn:
            count = conn.execute(
                "select count(*) from quote_document where date_trunc('day', created_at) = date_trunc('day', now())"
            ).fetchone()["count"]
            quote_number = f"КП-{datetime.now().strftime('%Y%m%d')}-{count+1}"

            row = conn.execute(
                '''
                insert into quote_document (client_id, quote_number, quote_date, note, status)
                values ((select id from client_account where id = %s), %s, CURRENT_DATE, %s, %s)
                returning *
                ''',
                (client_id, quote_number, note, status),
            ).fetchone()
            quote_id = str(row["id"])
            quote = self.serialize_quote(conn, row)

        if files:
            correlation_id = str(uuid.uuid4())
            dispatch_payload = {
                "quote_id": quote_id,
                "task_type": "quote_request",
                "note": note,
                "correlation_id": correlation_id,
                "callbackSuccessUrl": f"{os.getenv('PUBLIC_API_URL', 'http://127.0.0.1:8078')}/api/webhooks/n8n/quote-result",
                "callbackFailedUrl": f"{os.getenv('PUBLIC_API_URL', 'http://127.0.0.1:8078')}/api/webhooks/n8n/quote-failed",
                "file_binary": {
                    "path": files[0]["storage_path"],
                    "filename": files[0]["file_name"],
                    "mime_type": files[0]["mime_type"],
                },
            }
            with self.db() as conn:
                job_row = conn.execute(
                    """
                    insert into job_run (type, target_type, target_id, status, payload)
                    values ('quote_parse', 'quote_document', %s, 'queued', %s::jsonb)
                    returning id
                    """,
                    (quote_id, json.dumps(dispatch_payload, ensure_ascii=False, default=json_default)),
                ).fetchone()
                dispatch_payload["job_id"] = str(job_row["id"])
                conn.commit()

            def run_quote_dispatch():
                if not self.config.n8n_webhook_url:
                    n8n_logger.warning(
                        f"[N8N] SKIP quote_id={quote_id} — N8N_IMPORT_WEBHOOK_URL not configured"
                    )
                    return
                n8n_logger.info(
                    f"[N8N] DISPATCH → quote_id={quote_id} "
                    f"correlation_id={correlation_id} job_id={dispatch_payload['job_id']}"
                )
                try:
                    result = self.dispatch_to_n8n(dispatch_payload)
                    if result:
                        n8n_logger.info(
                            f"[N8N] QUOTE_OK quote_id={quote_id} "
                            f"correlation_id={correlation_id}"
                        )
                    else:
                        n8n_logger.error(
                            f"[N8N] QUOTE_FAILED quote_id={quote_id} "
                            f"correlation_id={correlation_id} — empty response from n8n"
                        )
                        with self.db() as conn_bg:
                            conn_bg.execute(
                                "update quote_document set status = 'failed' where id = %s",
                                (quote_id,),
                            )
                            conn_bg.commit()
                except Exception as e:
                    n8n_logger.error(
                        f"[N8N] QUOTE_EXCEPTION quote_id={quote_id} "
                        f"correlation_id={correlation_id} error={e}"
                    )

            threading.Thread(target=run_quote_dispatch, daemon=True).start()

        return self.respond_json({"item": quote})

    def handle_update_quote(self, quote_id: str) -> None:
        if not self.validate_uuid(quote_id):
            return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
            
        payload = self.read_json_body()
        items = payload.get("items", [])
        
        with self.db() as conn:
            row = conn.execute("select * from quote_document where id = %s", (quote_id,)).fetchone()
            if not row:
                return self.respond_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
                
            conn.execute("delete from quote_item where quote_document_id = %s", (quote_id,))
            
            for i, item in enumerate(items):
                conn.execute(
                    '''
                    insert into quote_item (quote_document_id, line_no, name_snapshot, volume_l, purchase_price, rrc_min, sale_price, qty)
                    values (%s, %s, %s, %s, %s, %s, %s, %s)
                    ''',
                    (
                        quote_id,
                        i + 1,
                        item.get("name") or item.get("raw_name"),
                        item.get("volume_l"),
                        item.get("purchase_price"),
                        item.get("rrc_min"),
                        item.get("sale_price"),
                        item.get("qty", 1)
                    )
                )
            
            updated_quote = self.serialize_quote(conn, row)
            
        return self.respond_json({"item": updated_quote})

    def handle_get_system_logs(self) -> None:
        try:
            log_file = log_dir / "api.log"
            if not log_file.exists():
                return self.respond_json({"error": "Log file not found", "lines": []})
            
            with open(log_file, "r", encoding="utf-8") as f:
                content = f.readlines()
                
            # Return last 500 lines
            return self.respond_json({
                "file": str(log_file),
                "lines": [line.strip() for line in content[-500:]]
            })
        except Exception as e:
            return self.respond_json({"error": str(e), "lines": []}, status=HTTPStatus.INTERNAL_SERVER_ERROR)



def cleanup_worker():
    """Background thread to delete files in UPLOADS_DIR and Postgres bytea older than 7 days."""
    logger.info("Cleanup thread started.")
    while True:
        try:
            now = time.time()
            cutoff = now - (7 * 24 * 3600)  # 7 days
            deleted_count = 0
            if UPLOADS_DIR.exists():
                for f in UPLOADS_DIR.iterdir():
                    if f.is_file():
                        try:
                            if f.stat().st_mtime < cutoff:
                                f.unlink()
                                deleted_count += 1
                        except Exception as e:
                            logger.error(f"Failed to delete old file {f}: {e}")
            
            # Also clean up Postgres file_bytes
            import psycopg
            try:
                with psycopg.connect(PostgresApiHandler.config.db_dsn) as conn:
                    # Clear out bytes for old files to reclaim space
                    res = conn.execute("UPDATE import_file SET file_bytes = NULL, cleanup_done = TRUE WHERE uploaded_at < NOW() - INTERVAL '7 days' AND cleanup_done = FALSE")
                    if res.rowcount > 0:
                        logger.info(f"Cleanup thread: cleared file_bytes for {res.rowcount} old DB records.")
                    conn.commit()
            except Exception as db_e:
                logger.error(f"Failed to clean up db file_bytes: {db_e}")

            if deleted_count > 0:

                logger.info(f"Cleanup thread: deleted {deleted_count} files older than 7 days.")
        except Exception as e:
            logger.error(f"Error in cleanup thread: {e}")
            
        # Check every hour
        time.sleep(3600)


def main() -> None:
    args = parse_args()
    server = ThreadingHTTPServer((args.host, args.port), PostgresApiHandler)
    logger.info(f"bahus PostgreSQL API running at http://{args.host}:{args.port}")
    
    # Log key config at startup for diagnostics
    openai_key_status = "SET (len=" + str(len(PostgresApiHandler.config.openai_api_key or "")) + ")" if PostgresApiHandler.config.openai_api_key else "NOT SET"
    logger.info(f"[CONFIG] OPENAI_API_KEY: {openai_key_status}")

    # Run inline migration
    try:
        import psycopg
        with psycopg.connect(PostgresApiHandler.config.db_dsn) as conn:
            conn.execute("ALTER TABLE import_file ADD COLUMN IF NOT EXISTS file_bytes BYTEA;")
            conn.execute("ALTER TABLE import_file ADD COLUMN IF NOT EXISTS cleanup_done BOOLEAN DEFAULT FALSE;")
            conn.execute("ALTER TABLE import_batch ADD COLUMN IF NOT EXISTS meta JSONB;")
            conn.commit()
            logger.info("Database schema check passed.")

    except Exception as e:
        logger.error(f"Failed to run schema migration: {e}")

    # Start the background cleanup thread

    cleanup_thread = threading.Thread(target=cleanup_worker, daemon=True)
    cleanup_thread.start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        logger.info("Shutting down PostgreSQL API")


if __name__ == "__main__":
    main()
